import openpyxl

wb = openpyxl.load_workbook("Book5.xlsx")
sheet = wb["Sheet1"]

template = openpyxl.load_workbook("test.xlsx")
temp_sheet = template["Sheet1"]


def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    for i in range(startRow, endRow + 1, 1):
        rowSelected = []
        for j in range(startCol, endCol + 1, 1):
            rowSelected.append(sheet.cell(row=i, column=j).value)
        rangeSelected.append(rowSelected)

    return rangeSelected


def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
    countRow = 0
    for i in range(startRow, endRow + 1, 1):
        countCol = 0
        for j in range(startCol, endCol + 1, 1):

            sheetReceiving.cell(row=i, column=j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1


def createData():
    print("Processing...")
    selectedRange = copyRange(2, 2, 2, 343, sheet)
    pastingRange = pasteRange(2, 2, 2, 343, temp_sheet, selectedRange)
    template.save("test.xlsx")
    print("Range copied and pasted!")


createData()
